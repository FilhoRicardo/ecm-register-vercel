from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "external" / "crrem" / "v2.05" / "emission-factors-v2.05.xlsx"
OUTPUT = ROOT / "src" / "data" / "crremEmissionFactorsV205.js"


def carrier_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def main() -> None:
    workbook_hash = hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb["Emission Factors"]

    years = [int(ws.cell(5, column).value) for column in range(4, 35)]
    grid = {}
    sources = {}
    for row in range(6, 90):
        country = ws.cell(row, 3).value
        if not country:
            continue
        values = [ws.cell(row, column).value for column in range(4, 35)]
        if all(value is None for value in values):
            continue
        grid[str(country)] = {
            str(year): float(value or 0)
            for year, value in zip(years, values)
        }
        sources[str(country)] = {
            "region": str(ws.cell(row, 1).value or ""),
            "code": str(ws.cell(row, 2).value or ""),
            "actual": str(ws.cell(row, 35).value or ""),
            "forecast": str(ws.cell(row, 36).value or ""),
            "notes": str(ws.cell(row, 37).value or ""),
        }

    fixed = {}
    for row in range(94, 102):
        carrier = ws.cell(row, 1).value
        if not carrier:
            continue
        factor = ws.cell(row, 2).value
        fixed[carrier_key(str(carrier))] = {
            "label": str(carrier),
            "factor": None if str(factor).lower() == "varies" else float(factor),
            "unit": str(ws.cell(row, 3).value or ""),
            "notes": str(ws.cell(row, 4).value or ""),
            "source": str(ws.cell(row, 35).value or ""),
        }

    text = "\n".join([
        "// Generated from CRREM Emission Factors workbook v2.05. Source: https://crrem.org/learn/",
        "// Workbook: external/crrem/v2.05/emission-factors-v2.05.xlsx",
        f"// Workbook SHA256: {workbook_hash}",
        'export const CRREM_EMISSION_FACTORS_VERSION = "v2.05";',
        'export const CRREM_EMISSION_FACTORS_SOURCE = "CRREM Foundation, Emission Factors v2.05, https://crrem.org/learn/";',
        f'export const CRREM_EMISSION_FACTORS_WORKBOOK_SHA256 = "{workbook_hash}";',
        f"export const CRREM_GRID_EF_OFFICIAL = {json.dumps(grid, ensure_ascii=False, separators=(',', ':'))};",
        f"export const CRREM_GRID_EF_SOURCES = {json.dumps(sources, ensure_ascii=False, separators=(',', ':'))};",
        f"export const CRREM_FIXED_EF_OFFICIAL = {json.dumps(fixed, ensure_ascii=False, separators=(',', ':'))};",
        "",
    ])
    OUTPUT.write_text(text, encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
